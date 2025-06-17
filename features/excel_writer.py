# features/excel_writer.py

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def write_resource_sheet(wb, resource_name, data):
    safe_title = resource_name[:31]
    ws = wb.create_sheet(title=safe_title)

    if not data:
        ws.append(["No data found."])
        return

    if not isinstance(data, list) or not isinstance(data[0], dict):
        ws.append(["Invalid data format."])
        return

    headers = list(data[0].keys())
    ws.append(headers)

    # Style setup
    header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    data_font = Font(size=10, name="Calibri")
    resource_id_font = Font(size=11, name="Calibri", bold=True)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", fill_type="solid")
    highlight_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Header formatting
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = max(22, len(header) + 5)

    # Data rows
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, key in enumerate(headers, start=1):
            value = row_data.get(key, "")
            if isinstance(value, dict):
                value = str(value)
            elif key == "Notes" and isinstance(value, list):
                value = ', '.join(value)
            elif isinstance(value, list):
                value = ', '.join(str(v) for v in value)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if key.lower() in ["resource id", "resource_id", "bucket name", "db instance identifier"]:
                cell.font = resource_id_font
            else:
                cell.font = data_font

            cell.alignment = alignment
            cell.border = thin_border

            if key.lower() in ["used?", "idle?", "underutilized?", "used in last 30 days"]:
                if str(value).lower() in ["no", "false"]:
                    cell.fill = highlight_fill

        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = alt_fill

def save_report(filename, resource_data_map):
    wb = Workbook()
    wb.remove(wb.active)

    for resource_name, data in resource_data_map.items():
        if resource_name == "RDS - Instances" and isinstance(data, dict):
            for sub_key, sub_data in data.items():
                if isinstance(sub_data, list) and sub_data:
                    sheet_title = f"RDS - {sub_key}"[:31]
                    write_resource_sheet(wb, sheet_title, sub_data)
        elif isinstance(data, list) and data:
            sheet_title = resource_name[:31]
            write_resource_sheet(wb, sheet_title, data)

    wb.save(filename)
